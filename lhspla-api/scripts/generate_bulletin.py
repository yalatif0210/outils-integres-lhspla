#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_bulletin.py
Génère le Bulletin Hebdomadaire LHSPLA-TA depuis PostgreSQL (openpyxl).

Usage:
  python generate_bulletin.py [--semaine YYYY-MM-DD] [--output-dir DIR]

Le fichier de référence doit se trouver dans le même répertoire que ce script.
"""
import argparse
import os
import shutil
import sys
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─── Configuration ─────────────────────────────────────────────────────────────

DB_DSN = (
    "host=localhost port=5432 dbname=lhspla_bulletin "
    "user=postgres password=7d12b33499a149589f605989f1332188"
)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REFERENCE_FILE = os.path.join(
    SCRIPT_DIR,
    "Bulletin Hebdo LHSPLA_V02.05.2026_du 25 au 29 mai 2026.xlsx",
)

# Ordre des entités dans COMPILATION et BulletinDu
ENTITY_ORDER = ["CAD", "PMO", "QAD", "SE", "SI", "FINANCES", "CAC", "COM"]

FRENCH_MONTHS = {
    1: "janvier", 2: "février", 3: "mars", 4: "avril",
    5: "mai", 6: "juin", 7: "juillet", 8: "août",
    9: "septembre", 10: "octobre", 11: "novembre", 12: "décembre",
}

CRITICALITY_DISPLAY = {
    "critique": "🔴 Critique",
    "eleve":    "🟠 Élevé",
    "modere":   "🟡 Modéré",
    "faible":   "🟢 Faible",
}
DOS_DISPLAY = {"oui": "✅ OUI", "non": "❌ NON"}

# ─── Couleurs ARGB (alpha FF = opaque) ────────────────────────────────────────

C_DARK_BLUE  = "FF1F4E79"   # en-tête principal
C_MED_BLUE   = "FF2E75B6"   # sous-en-tête entité / soumission
C_ORANGE     = "FFC55A11"   # bannière Section B
C_GREEN      = "FF548235"   # bannière Section C
C_DARK_RED   = "FF9E0000"   # bannière Section D
C_LIGHT_BLUE = "FFBDD7EE"   # en-têtes colonnes (toutes sections)
C_PALE_BLUE  = "FFDEEAF1"   # en-têtes soumission (Section A)
C_YELLOW     = "FFFFF2CC"   # cellules de saisie
C_GRAY_N     = "FFF2F2F2"   # cellule N°
C_GRAY_CODE  = "FFD9D9D9"   # code entité
C_NOTE_B     = "FFFCE4D6"   # note / en-tête entité COMPILATION B
C_NOTE_C     = "FFE2EFDA"   # note / en-tête entité COMPILATION C
C_NOTE_D     = "FFFFE0E0"   # note / en-tête entité COMPILATION D
C_WHITE      = "FFFFFFFF"


# ─── Helpers de mise en forme ─────────────────────────────────────────────────

def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _font(bold=False, italic=False, size=9, color="FF000000") -> Font:
    return Font(bold=bold, italic=italic, size=size, color=color)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _w(ws, row, col, value, f=None, fn=None, al=None):
    """Écrit une cellule avec mise en forme."""
    c = ws.cell(row=row, column=col, value=value)
    if f:  c.fill = f
    if fn: c.font = fn
    if al: c.alignment = al
    return c


def _mw(ws, row, col1, col2, value, f=None, fn=None, al=None):
    """Fusionne col1:col2 sur row puis écrit la valeur."""
    ws.merge_cells(
        start_row=row, start_column=col1,
        end_row=row,   end_column=col2,
    )
    return _w(ws, row, col1, value, f, fn, al)


_THIN = Side(style="thin")


def _b(t=False, bo=False, l=False, r=False) -> Border:
    return Border(
        top=_THIN if t else None,
        bottom=_THIN if bo else None,
        left=_THIN if l else None,
        right=_THIN if r else None,
    )


def _apply_entity_borders(ws):
    """Applique les bordures sur toutes les cellules de la feuille entité."""
    for row in range(1, 44):
        if row in (3, 7, 19, 20, 35):
            continue

        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)

            if row in (1, 4, 8, 9, 21, 22, 36, 37):
                # Ligne entière fusionnée A:H → L seulement col1
                cell.border = _b(t=True, bo=True, l=(col == 1))

            elif row == 2:
                # Deux fusions : A:E et F:H
                cell.border = _b(t=True, bo=True, l=(col in (1, 6)))

            elif row == 5:
                # E5/F5 sont maintenant individuels (Date de début / Date de fin)
                if col in (4, 8):
                    cell.border = _b(t=True, bo=True, r=True)
                else:
                    cell.border = _b(t=True, bo=True, l=True, r=True)

            elif row == 6:
                # F6 est maintenant individuel (Date de fin) → ajouter L
                if col == 1:
                    cell.border = _b(t=True, bo=True, l=True, r=True)
                elif col in (2, 3, 5, 6, 7):
                    cell.border = _b(t=True, bo=True, l=True)
                else:  # col in (4, 8)
                    cell.border = _b(t=True, bo=True)

            elif row in range(10, 19):
                # Section B header + data : 8 colonnes individuelles (plus de fusion G:H)
                cell.border = _b(t=True, bo=True, l=True, r=True)

            elif row in range(23, 34):
                # Section C header + data : toutes colonnes individuelles
                cell.border = _b(t=True, bo=True, l=True, r=True)

            elif row == 38:
                # Section D header (G:H fusionné, même pattern que section B)
                if col == 8:
                    cell.border = _b(t=True, bo=True, r=True)
                else:
                    cell.border = _b(t=True, bo=True, l=True, r=True)

            elif row in range(39, 44):
                # Section D data (G:H fusionné) : G→TBoL, H→TBo
                if col == 7:
                    cell.border = _b(t=True, bo=True, l=True)
                elif col == 8:
                    cell.border = _b(t=True, bo=True)
                else:
                    cell.border = _b(t=True, bo=True, l=True, r=True)


def _fmt_date(dt) -> str:
    """Formate une date en JJ/MM/AAAA, retourne '' si None."""
    if dt is None:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%d/%m/%Y")
    parts = str(dt)[:10].split("-")
    return f"{parts[2]}/{parts[1]}/{parts[0]}" if len(parts) == 3 else str(dt)


def _apply_comp_borders(ws):
    """Bordures pour COMPILATION et BulletinDu (7 colonnes)."""
    # R1 : en-tête fusion A:G — L seulement col1
    for col in range(1, 8):
        ws.cell(1, col).border = _b(t=True, bo=True, l=(col == 1))
    # R2 : deux fusions A:D et E:G
    for col in range(1, 8):
        ws.cell(2, col).border = _b(t=True, bo=True, l=(col in (1, 5)))

    # Identifier lignes pleine-largeur fusionnées (bannières + noms entités)
    full_spans = set()
    for m in ws.merged_cells.ranges:
        if m.min_col == 1 and m.max_col == 7:
            for row in range(m.min_row, m.max_row + 1):
                full_spans.add(row)

    # Lignes espaceur (hauteur == 8)
    spacers = {3}
    for row in range(4, ws.max_row + 1):
        if ws.row_dimensions[row].height == 8:
            spacers.add(row)

    for row in range(4, ws.max_row + 1):
        if row in spacers:
            continue
        merged = row in full_spans
        ws.cell(row, 1).border = _b(t=True, bo=True, l=True, r=True)
        for col in range(2, 7):
            if merged:
                ws.cell(row, col).border = _b(t=True, bo=True)
            else:
                ws.cell(row, col).border = _b(t=True, bo=True, l=True, r=True)
        ws.cell(row, 7).border = _b(t=True, bo=True, r=True)


def _reset_sheet(wb, name: str):
    """Supprime l'onglet et le recrée vierge à la même position."""
    names = wb.sheetnames
    idx = names.index(name) if name in names else len(names)
    if name in names:
        del wb[name]
    return wb.create_sheet(name, idx)


# ─── Base de données ───────────────────────────────────────────────────────────

def db_connect():
    return psycopg2.connect(DB_DSN, cursor_factory=psycopg2.extras.RealDictCursor)


def load_week(conn, semaine: str | None) -> dict:
    with conn.cursor() as cur:
        if semaine:
            cur.execute(
                """SELECT id, "weekReference",
                          "weekStart"::date::text AS "weekStart",
                          "weekEnd"::date::text   AS "weekEnd"
                   FROM "Week"
                   WHERE "weekStart"::date = %s::date
                   LIMIT 1""",
                (semaine,),
            )
        else:
            cur.execute(
                """SELECT id, "weekReference",
                          "weekStart"::date::text AS "weekStart",
                          "weekEnd"::date::text   AS "weekEnd"
                   FROM "Week"
                   ORDER BY "weekStart" DESC LIMIT 1"""
            )
        row = cur.fetchone()
    if not row:
        sys.exit("❌  Aucune semaine trouvée en base de données.")
    return dict(row)


def load_entity_data(conn, week_id: str) -> dict:
    """Retourne {entityCode: {submission, activities, planned, risks}}."""
    with conn.cursor() as cur:
        cur.execute(
            """SELECT id, "entityCode", responsible, "submissionDate"
               FROM "EntitySubmission" WHERE "weekId" = %s""",
            (week_id,),
        )
        submissions = {r["entityCode"]: dict(r) for r in cur.fetchall()}

        sub_ids = [s["id"] for s in submissions.values()] if submissions else []

        acts_by_sub: dict = {}
        plan_by_sub: dict = {}
        risk_by_sub: dict = {}

        if sub_ids:
            cur.execute(
                """SELECT "submissionId", "orderIndex",
                          title, objectives, location, dates, recommendations,
                          "startDate", "endDate"
                   FROM "Activity"
                   WHERE "submissionId" = ANY(%s)
                   ORDER BY "submissionId", "orderIndex" """,
                (sub_ids,),
            )
            for r in cur.fetchall():
                acts_by_sub.setdefault(r["submissionId"], []).append(dict(r))

            cur.execute(
                """SELECT "submissionId", "orderIndex",
                          title, objectives, location,
                          "plannedDates", "startDate", "endDate",
                          "dosParticipation", observations
                   FROM "PlannedActivity"
                   WHERE "submissionId" = ANY(%s)
                   ORDER BY "submissionId", "orderIndex" """,
                (sub_ids,),
            )
            for r in cur.fetchall():
                plan_by_sub.setdefault(r["submissionId"], []).append(dict(r))

            cur.execute(
                """SELECT "submissionId", "orderIndex",
                          theme, category, description,
                          criticality, "expectedAction"
                   FROM "RiskPoint"
                   WHERE "submissionId" = ANY(%s)
                   ORDER BY "submissionId", "orderIndex" """,
                (sub_ids,),
            )
            for r in cur.fetchall():
                risk_by_sub.setdefault(r["submissionId"], []).append(dict(r))

    result = {}
    for code in ENTITY_ORDER:
        sub = submissions.get(code)
        sid = sub["id"] if sub else None
        result[code] = {
            "submission": sub,
            "activities": acts_by_sub.get(sid, []) if sid else [],
            "planned":    plan_by_sub.get(sid, []) if sid else [],
            "risks":      risk_by_sub.get(sid, []) if sid else [],
        }
    return result


# ─── Construction : onglet entité ─────────────────────────────────────────────
#
# Structure standard (8 colonnes A-H) :
#   Row 1        : En-tête principal           A1:H1
#   Row 2        : Nom entité A2:E2 + version F2:H2
#   Row 3        : vide
#   Row 4        : Section A  A4:H4
#   Row 5        : Sous-en-têtes soumission    C5:D5 | E5:F5 | G5:H5
#   Row 6        : Données soumission          A6:B6 | C6:D6 | E6:F6 | G6:H6
#   Row 7        : vide
#   Row 8        : Bannière Section B          A8:H8
#   Row 9        : Note Section B              A9:H9
#   Row 10       : En-têtes colonnes Section B (G10:H10 fusionné)
#   Rows 11-18   : 8 activités réalisées       (G:H fusionné par ligne)
#   Row 19       : vide fusionné               A19:H19
#   Row 20       : vide
#   Row 21       : Bannière Section C          A21:H21
#   Row 22       : Note Section C              A22:H22
#   Row 23       : En-têtes colonnes Section C (8 colonnes individuelles)
#   Rows 24-33   : 10 activités planifiées     (toutes colonnes individuelles)
#   Row 34       : Note DoS                    A34:H34
#   Row 35       : vide
#   Row 36       : Bannière Section D          A36:H36
#   Row 37       : Note Section D              A37:H37
#   Row 38       : En-têtes colonnes Section D (G38:H38 fusionné vide)
#   Rows 39-43   : 5 points de vigilance       (G:H fusionné vide par ligne)
# ─────────────────────────────────────────────────────────────────────────────

ENTITY_COL_WIDTHS = {
    "A": 4.5703125, "B": 28.5703125, "C": 30.0, "D": 32.0,
    "E": 13.0,      "F": 14.0,       "G": 20.0, "H": 30.0,
}


def build_entity_sheet(ws, code: str, full_name: str, data: dict, week_ref: str,
                       week_start: str = "", week_end: str = "",
                       col_widths: dict | None = None):
    sub    = data["submission"] or {}
    acts   = data["activities"]
    plannd = data["planned"]
    risks  = data["risks"]

    # Priorité aux largeurs lues depuis la référence ; fallback sur défauts
    widths = col_widths if col_widths else ENTITY_COL_WIDTHS
    for col, w in widths.items():
        if w:
            ws.column_dimensions[col].width = w

    # Styles
    fw13  = _font(bold=True, size=13, color=C_WHITE)
    fw10  = _font(bold=True, size=10, color=C_WHITE)
    fw9i  = _font(italic=True, size=9, color=C_WHITE)
    fb9   = _font(bold=True, size=9)
    f9    = _font(size=9)
    f9i   = _font(italic=True, size=9)
    ac    = _align("center", "center")
    ac_w  = _align("center", "center", wrap=True)
    al_w  = _align("left", "center", wrap=True)

    al     = _align("left", "center")
    ar     = _align("right", "center")

    # ── Ligne 1 ──────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30.0
    _mw(ws, 1, 1, 8,
        "BULLETIN HEBDOMADAIRE LHSPLA-TA  —  CANEVAS DE REPORTING ENTITÉ",
        _fill(C_DARK_BLUE), fw13, ac)

    # ── Ligne 2 ──────────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 21.75
    _mw(ws, 2, 1, 5, full_name, _fill(C_MED_BLUE), fw10, al)
    c6 = _mw(ws, 2, 6, 8, "", _fill(C_MED_BLUE), fw9i, _align("right", "center"))

    # ── Ligne 3 : vide (hauteur par défaut) ──────────────────────────────────

    # ── Ligne 4 : Section A ──────────────────────────────────────────────────
    ws.row_dimensions[4].height = 19.5
    _mw(ws, 4, 1, 8, "  A.  INFORMATIONS DE SOUMISSION",
        _fill(C_MED_BLUE), fw10, al)

    # ── Ligne 5 : sous-en-têtes ──────────────────────────────────────────────
    ws.row_dimensions[5].height = 18.0
    entity_label = "Entité / Unité" if code == "FINANCES" else "Entité"
    _w(ws, 5, 1, entity_label,               _fill(C_PALE_BLUE), fb9, ac)
    _w(ws, 5, 2, "",                          _fill(C_PALE_BLUE), fb9, ac)
    _mw(ws, 5, 3, 4, "Responsable (Nom & Prénom)", _fill(C_PALE_BLUE), fb9, ac)
    _w(ws, 5, 5, "Date de début",            _fill(C_PALE_BLUE), fb9, ac)
    _w(ws, 5, 6, "Date de fin",              _fill(C_PALE_BLUE), fb9, ac)
    _mw(ws, 5, 7, 8, "Date de soumission",   _fill(C_PALE_BLUE), fb9, ac)

    # ── Ligne 6 : données soumission ─────────────────────────────────────────
    ws.row_dimensions[6].height = 21.75
    _w(ws, 6, 1, code, _fill(C_GRAY_CODE), fb9, ac)
    _w(ws, 6, 2, "",   _fill(C_GRAY_CODE))
    _mw(ws, 6, 3, 4, sub.get("responsible", ""), _fill(C_YELLOW), f9, ac_w)
    if code in ("SE", "FINANCES"):
        _w(ws, 6, 5, _fmt_date(week_start),              _fill(C_YELLOW), f9, ac)
        _w(ws, 6, 6, _fmt_date(week_end),                _fill(C_YELLOW), f9, ac)
        _mw(ws, 6, 7, 8, sub.get("submissionDate", ""), _fill(C_YELLOW), f9, ac)
    else:
        _w(ws, 6, 5, "=SE!E6",  _fill(C_YELLOW), f9, ac)
        _w(ws, 6, 6, "=SE!F6",  _fill(C_YELLOW), f9, ac)
        _mw(ws, 6, 7, 8, "=SE!G6", _fill(C_YELLOW), f9, ac)

    # ── Ligne 7 : vide (hauteur par défaut) ──────────────────────────────────

    # ── Ligne 8 : bannière Section B ─────────────────────────────────────────
    ws.row_dimensions[8].height = 21.75
    _mw(ws, 8, 1, 8,
        "  B.  ACTIVITÉS RÉALISÉES  —  Semaine précédente",
        _fill(C_ORANGE), fw10, al)

    # ── Ligne 9 : note Section B ─────────────────────────────────────────────
    ws.row_dimensions[9].height = 15.75
    _mw(ws, 9, 1, 8,
        "ℹ️  Pour chaque activité réalisée au cours de la semaine écoulée, compléter une ligne.",
        _fill(C_NOTE_B), f9i, al)

    # ── Ligne 10 : en-têtes colonnes Section B ───────────────────────────────
    ws.row_dimensions[10].height = 30.0
    b_hdrs = [
        "N°", "N° Activités", "Libellé de la tâche", "Objectifs", "Lieu",
        "Date de début", "Date de fin",
        "Principales recommandations / Résultats clés",
    ]
    for i, txt in enumerate(b_hdrs, start=1):
        _w(ws, 10, i, txt, _fill(C_LIGHT_BLUE), fb9, ac_w)

    # ── Lignes 11-18 : activités réalisées ───────────────────────────────────
    for i in range(8):
        row = 11 + i
        act = acts[i] if i < len(acts) else {}
        _w(ws, row, 1, i + 1,                           _fill(C_GRAY_N),  fb9, ac)
        _w(ws, row, 2, act.get("orderIndex", ""),        _fill(C_YELLOW),  f9,  ac)
        _w(ws, row, 3, act.get("title", ""),             _fill(C_YELLOW),  f9,  al_w)
        _w(ws, row, 4, act.get("objectives", ""),        _fill(C_YELLOW),  f9,  al_w)
        _w(ws, row, 5, act.get("location", ""),          _fill(C_YELLOW),  f9,  al_w)
        _w(ws, row, 6, _fmt_date(act.get("startDate")),     _fill(C_YELLOW), f9, ac_w)
        _w(ws, row, 7, _fmt_date(act.get("endDate")),       _fill(C_YELLOW), f9, ac_w)
        _w(ws, row, 8, act.get("recommendations", ""),      _fill(C_YELLOW), f9, al_w)

    # ── Ligne 19 : note ajout lignes ─────────────────────────────────────────
    ws.row_dimensions[19].height = 12.75
    _mw(ws, 19, 1, 8,
        "↳  Ajouter autant de lignes que nécessaire avant soumission.",
        None, _font(italic=True, size=8, color="FF595959"), ar)

    # ── Ligne 20 : vide (hauteur par défaut) ─────────────────────────────────

    # ── Ligne 21 : bannière Section C ────────────────────────────────────────
    ws.row_dimensions[21].height = 21.75
    _mw(ws, 21, 1, 8,
        "  C.  ACTIVITÉS PLANIFIÉES  —  Mois glissant (4 semaines à venir)",
        _fill(C_GREEN), fw10, al)

    # ── Ligne 22 : note Section C ────────────────────────────────────────────
    ws.row_dimensions[22].height = 15.75
    _mw(ws, 22, 1, 8,
        "ℹ️  Lister toutes les activités prévues sur les 4 prochaines semaines."
        " Signaler en ROUGE les activités nécessitant la participation du DoS.",
        _fill(C_NOTE_C), f9i, al)

    # ── Ligne 23 : en-têtes colonnes Section C ───────────────────────────────
    ws.row_dimensions[23].height = 30.0
    c_hdrs = [
        "N°", "N° Activités", "Libellé de la tâche", "Objectifs",
        "Lieu", "Date(s) prévues", "Participation DoS ?", "Observations / Points d'attention",
    ]
    for i, txt in enumerate(c_hdrs, start=1):
        _w(ws, 23, i, txt, _fill(C_LIGHT_BLUE), fb9, ac_w)

    # ── Lignes 24-33 : activités planifiées ──────────────────────────────────
    for i in range(10):
        row = 24 + i
        pl       = plannd[i] if i < len(plannd) else {}
        dos_raw  = pl.get("dosParticipation") or ""
        _w(ws, row, 1, i + 1,                            _fill(C_GRAY_N), fb9, ac)
        _w(ws, row, 2, pl.get("orderIndex", ""),          _fill(C_YELLOW), f9,  ac)
        _w(ws, row, 3, pl.get("title", ""),               _fill(C_YELLOW), f9,  al_w)
        _w(ws, row, 4, pl.get("objectives", ""),          _fill(C_YELLOW), f9,  al_w)
        _w(ws, row, 5, pl.get("location", ""),            _fill(C_YELLOW), f9,  al_w)
        _w(ws, row, 6, pl.get("plannedDates", ""),        _fill(C_YELLOW), f9,  ac_w)
        _w(ws, row, 7, DOS_DISPLAY.get(dos_raw, dos_raw), _fill(C_YELLOW), f9,  ac)
        _w(ws, row, 8, pl.get("observations", ""),        _fill(C_YELLOW), f9,  al_w)

    # ── Ligne 34 : note DoS ──────────────────────────────────────────────────
    ws.row_dimensions[34].height = 12.75
    _mw(ws, 34, 1, 8,
        "↳  Toute activité impliquant le DoS doit être transmise au moins 2 semaines avant son démarrage.",
        None, _font(italic=True, size=8, color="FF595959"), ar)

    # ── Ligne 35 : vide (hauteur par défaut) ─────────────────────────────────

    # ── Ligne 36 : bannière Section D ────────────────────────────────────────
    ws.row_dimensions[36].height = 21.75
    _mw(ws, 36, 1, 8,
        "  D.  POINTS DE VIGILANCE  /  RISQUES  /  BESOINS DE DÉCISION",
        _fill(C_DARK_RED), fw10, al)

    # ── Ligne 37 : note Section D ────────────────────────────────────────────
    ws.row_dimensions[37].height = 15.75
    _mw(ws, 37, 1, 8,
        "ℹ️  Signaler tout élément nécessitant une attention particulière, "
        "une décision ou un arbitrage.",
        _fill(C_NOTE_D), f9i, al)

    # ── Ligne 38 : en-têtes colonnes Section D ───────────────────────────────
    ws.row_dimensions[38].height = 30.0
    d_hdrs = ["N°", "Thème", "Catégorie", "Description du point", "Niveau de criticité",
              "Action attendue / Décision souhaitée"]
    for i, txt in enumerate(d_hdrs, start=1):
        _w(ws, 38, i, txt, _fill(C_LIGHT_BLUE), fb9, ac_w)
    ws.merge_cells(start_row=38, start_column=7, end_row=38, end_column=8)
    _w(ws, 38, 7, "", _fill(C_LIGHT_BLUE), fb9, ac_w)

    # ── Lignes 39-43 : points de vigilance ───────────────────────────────────
    for i in range(5):
        row      = 39 + i
        rk       = risks[i] if i < len(risks) else {}
        crit_raw = rk.get("criticality") or ""
        _w(ws, row, 1, i + 1,                                    _fill(C_GRAY_N), fb9, ac)
        _w(ws, row, 2, rk.get("theme", ""),                      _fill(C_YELLOW), f9,  al_w)
        _w(ws, row, 3, rk.get("category", ""),                   _fill(C_YELLOW), f9,  al_w)
        _w(ws, row, 4, rk.get("description", ""),                _fill(C_YELLOW), f9,  al_w)
        _w(ws, row, 5, CRITICALITY_DISPLAY.get(crit_raw, crit_raw), _fill(C_YELLOW), f9, ac)
        _w(ws, row, 6, rk.get("expectedAction", ""),             _fill(C_YELLOW), f9,  al_w)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        _w(ws, row, 7, "", _fill(C_GRAY_N))

    # ── Bordures (appliquées en dernier pour couvrir tous les éléments) ───────
    _apply_entity_borders(ws)


# ─── Construction : onglet COMPILATION ────────────────────────────────────────
#
# 7 colonnes (A-G)  |  Largeurs identiques à COMPILATION du fichier de référence
#
# Section B  : COMPILATION cols A-F  (6 cols)
#   formules : ENTITY!C{11-18} → D → E → F → G  →  COMPILATION B-F
# Section C  : COMPILATION cols A-G  (7 cols)
#   formules : ENTITY!C{24-33} → D → E → F → G → H  →  COMPILATION B-G
# Section D  : COMPILATION cols A-E  (5 cols)
#   formules : ENTITY!C{39-43} → D → E → F  →  COMPILATION B-E
# ─────────────────────────────────────────────────────────────────────────────

COMP_COL_WIDTHS = {
    "A": 2.86, "B": 30.0, "C": 32.0, "D": 23.29,
    "E": 14.0, "F": 20.0, "G": 30.0,
}


def build_compilation_sheet(ws, week_ref: str, entity_names: dict):
    for col, w in COMP_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    fw13  = _font(bold=True, size=13, color=C_WHITE)
    fw10  = _font(bold=True, size=10, color=C_WHITE)
    fb9   = _font(bold=True, size=9)
    f9    = _font(size=9)
    ac    = _align("center", "center")
    ac_w  = _align("center", "center", wrap=True)
    al_w  = _align("left", "center", wrap=True)

    cur = 1  # curseur de ligne

    # ── Ligne 1 : en-tête ───────────────────────────────────────────────────
    ws.row_dimensions[cur].height = 26
    _mw(ws, cur, 1, 7,
        "COMPILATION HEBDOMADAIRE  —  BULLETIN LHSPLA-TA",
        _fill(C_DARK_BLUE), fw13, ac)
    cur += 1

    # ── Ligne 2 : semaine ───────────────────────────────────────────────────
    ws.row_dimensions[cur].height = 18
    _mw(ws, cur, 1, 4, "Semaine de référence : (se renseigne automatiquement à partir des onglets entités)",
        _fill(C_MED_BLUE), _font(size=9, color=C_WHITE), _align("left", "center"))
    _mw(ws, cur, 5, 7, week_ref,
        _fill(C_MED_BLUE), _font(bold=True, size=10, color=C_WHITE), _align("left", "center"))
    cur += 1

    # ── Ligne 3 : vide ──────────────────────────────────────────────────────
    ws.row_dimensions[cur].height = 8
    cur += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION B
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[cur].height = 21
    _mw(ws, cur, 1, 7,
        "  B.  ACTIVITÉS RÉALISÉES  —  Semaine précédente  (toutes entités)",
        _fill(C_ORANGE), fw10, _align("left", "center"))
    cur += 1

    b_hdrs = [
        "N°", "Titre de l'activité", "Objectifs", "Lieu",
        "Date de début", "Date de fin",
        "Principales recommandations / Résultats clés",
    ]

    for code in ENTITY_ORDER:
        name = entity_names.get(code, code)

        # En-tête entité (span A-G)
        ws.row_dimensions[cur].height = 16
        _mw(ws, cur, 1, 7, name, _fill(C_NOTE_B), fb9, _align("center", "center"))
        cur += 1

        # En-têtes colonnes (7 cols)
        ws.row_dimensions[cur].height = 26
        for i, txt in enumerate(b_hdrs, start=1):
            _w(ws, cur, i, txt, _fill(C_PALE_BLUE), fb9, ac_w)
        cur += 1

        # 8 lignes de formules (entity rows 11-18, entity cols C-H → COMPILATION B-G)
        for j in range(8):
            er = 11 + j   # entity row
            ws.row_dimensions[cur].height = 35
            _w(ws, cur, 1, j + 1, _fill(C_GRAY_N), fb9, ac)
            for comp_col, ec in enumerate("CDEFGH", start=2):
                fml = f'=IF({code}!{ec}{er}="","",{code}!{ec}{er})'
                _w(ws, cur, comp_col, fml, None, f9, al_w)
            cur += 1

        # Ligne vide séparatrice
        ws.row_dimensions[cur].height = 8
        cur += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION C
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[cur].height = 21
    _mw(ws, cur, 1, 7,
        "  C.  ACTIVITÉS PLANIFIÉES  —  Mois glissant  (toutes entités)",
        _fill(C_GREEN), fw10, _align("left", "center"))
    cur += 1

    c_hdrs = [
        "N°", "Titre de l'activité", "Objectifs", "Lieu",
        "Date(s) prévues", "Participation DoS ?", "Observations / Points d'attention",
    ]

    for code in ENTITY_ORDER:
        name = entity_names.get(code, code)

        ws.row_dimensions[cur].height = 16
        _mw(ws, cur, 1, 7, name, _fill(C_NOTE_C), fb9, ac)
        cur += 1

        ws.row_dimensions[cur].height = 26
        for i, txt in enumerate(c_hdrs, start=1):
            _w(ws, cur, i, txt, _fill(C_PALE_BLUE), fb9, ac_w)
        cur += 1

        # 10 lignes de formules (entity rows 24-33, entity cols C-H → COMPILATION B-G)
        for j in range(10):
            er = 24 + j
            ws.row_dimensions[cur].height = 32
            _w(ws, cur, 1, j + 1, _fill(C_GRAY_N), fb9, ac)
            for comp_col, ec in enumerate("CDEFGH", start=2):
                fml = f'=IF({code}!{ec}{er}="","",{code}!{ec}{er})'
                _w(ws, cur, comp_col, fml, None, f9, al_w)
            cur += 1

        ws.row_dimensions[cur].height = 8
        cur += 1

    # ════════════════════════════════════════════════════════════════════════
    # SECTION D
    # ════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[cur].height = 21
    _mw(ws, cur, 1, 7,
        "  D.  POINTS DE VIGILANCE  /  RISQUES  /  BESOINS DE DÉCISION  (toutes entités)",
        _fill(C_DARK_RED), fw10, _align("left", "center"))
    cur += 1

    d_hdrs = [
        "N°", "Catégorie", "Description du point",
        "Niveau de criticité", "Action attendue / Décision souhaitée",
    ]

    for code in ENTITY_ORDER:
        name = entity_names.get(code, code)

        ws.row_dimensions[cur].height = 16
        _mw(ws, cur, 1, 7, name, _fill(C_NOTE_D), fb9, ac)
        cur += 1

        ws.row_dimensions[cur].height = 26
        for i, txt in enumerate(d_hdrs, start=1):
            _w(ws, cur, i, txt, _fill(C_PALE_BLUE), fb9, ac_w)
        cur += 1

        # 5 lignes de formules (entity rows 39-43, entity cols C-F → COMPILATION B-E)
        # Thème (entity col B) est volontairement ignoré
        for j in range(5):
            er = 39 + j
            ws.row_dimensions[cur].height = 35
            _w(ws, cur, 1, j + 1, _fill(C_GRAY_N), fb9, ac)
            for comp_col, ec in enumerate("CDEF", start=2):
                fml = f'=IF({code}!{ec}{er}="","",{code}!{ec}{er})'
                _w(ws, cur, comp_col, fml, None, f9, al_w)
            cur += 1

        ws.row_dimensions[cur].height = 8
        cur += 1

    _apply_comp_borders(ws)


# ─── Construction : onglet BulletinDu (valeurs réelles, dynamique) ────────────

def build_bulletin_sheet(ws, week_ref: str, entity_names: dict, all_data: dict):
    """Onglet de synthèse avec valeurs réelles, sans formules."""
    BULLETIN_COL_WIDTHS = {**COMP_COL_WIDTHS, "E": 16.0}
    for col, w in BULLETIN_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    fw13  = _font(bold=True, size=13, color=C_WHITE)
    fw10  = _font(bold=True, size=10, color=C_WHITE)
    fb9   = _font(bold=True, size=9)
    f9    = _font(size=9)
    ac    = _align("center", "center")
    ac_w  = _align("center", "center", wrap=True)
    al_w  = _align("left", "center", wrap=True)

    cur = 1

    # ── En-tête ──────────────────────────────────────────────────────────────
    ws.row_dimensions[cur].height = 26
    _mw(ws, cur, 1, 7,
        "COMPILATION HEBDOMADAIRE  —  BULLETIN LHSPLA-TA",
        _fill(C_DARK_BLUE), fw13, ac)
    cur += 1

    ws.row_dimensions[cur].height = 18
    _mw(ws, cur, 1, 4,
        "Semaine de référence : (se renseigne automatiquement à partir des onglets entités)",
        _fill(C_MED_BLUE), _font(size=9, color=C_WHITE), _align("left", "center"))
    _mw(ws, cur, 5, 7, week_ref,
        _fill(C_MED_BLUE), _font(bold=True, size=10, color=C_WHITE),
        _align("left", "center"))
    cur += 1

    ws.row_dimensions[cur].height = 8
    cur += 1

    # ══════════════════════════════════════════════
    # SECTION B
    # ══════════════════════════════════════════════
    ws.row_dimensions[cur].height = 21
    _mw(ws, cur, 1, 7,
        "  B.  ACTIVITÉS RÉALISÉES  —  Semaine précédente",
        _fill(C_ORANGE), fw10, _align("left", "center"))
    cur += 1

    b_hdrs = [
        "N°", "Titre de l'activité", "Objectifs", "Lieu",
        "Date de début", "Date de fin",
        "Principales recommandations / Résultats clés",
    ]

    has_b = False
    for code in ENTITY_ORDER:
        acts = [a for a in all_data[code]["activities"] if a.get("title", "").strip()]
        if not acts:
            continue
        has_b = True
        name = entity_names.get(code, code)

        ws.row_dimensions[cur].height = 16
        _mw(ws, cur, 1, 7, name, _fill(C_NOTE_B), fb9, ac)
        cur += 1

        ws.row_dimensions[cur].height = 26
        for i, txt in enumerate(b_hdrs, start=1):
            _w(ws, cur, i, txt, _fill(C_PALE_BLUE), fb9, ac_w)
        cur += 1

        for idx, act in enumerate(acts):
            ws.row_dimensions[cur].height = 40
            _w(ws, cur, 1, idx + 1,                              _fill(C_GRAY_N), fb9, ac)
            _w(ws, cur, 2, act.get("title", ""),                  None, f9, al_w)
            _w(ws, cur, 3, act.get("objectives", ""),             None, f9, al_w)
            _w(ws, cur, 4, act.get("location", ""),               None, f9, al_w)
            _w(ws, cur, 5, _fmt_date(act.get("startDate")),       None, f9, ac_w)
            _w(ws, cur, 6, _fmt_date(act.get("endDate")),         None, f9, ac_w)
            _w(ws, cur, 7, act.get("recommendations", ""),        None, f9, al_w)
            cur += 1

        ws.row_dimensions[cur].height = 8
        cur += 1

    if not has_b:
        ws.row_dimensions[cur].height = 20
        _mw(ws, cur, 1, 7, "(Aucune activité réalisée pour cette semaine)",
            None, _font(italic=True, size=9), _align("center", "center"))
        cur += 1

    # ══════════════════════════════════════════════
    # SECTION C
    # ══════════════════════════════════════════════
    ws.row_dimensions[cur].height = 21
    _mw(ws, cur, 1, 7,
        "  C.  ACTIVITÉS PLANIFIÉES  —  Mois glissant",
        _fill(C_GREEN), fw10, _align("left", "center"))
    cur += 1

    c_hdrs = [
        "N°", "Titre de l'activité", "Objectifs", "Lieu",
        "Date(s) prévues", "Participation DoS ?", "Observations / Points d'attention",
    ]

    has_c = False
    for code in ENTITY_ORDER:
        plans = [p for p in all_data[code]["planned"] if p.get("title", "").strip()]
        if not plans:
            continue
        has_c = True
        name = entity_names.get(code, code)

        ws.row_dimensions[cur].height = 16
        _mw(ws, cur, 1, 7, name, _fill(C_NOTE_C), fb9, ac)
        cur += 1

        ws.row_dimensions[cur].height = 26
        for i, txt in enumerate(c_hdrs, start=1):
            _w(ws, cur, i, txt, _fill(C_PALE_BLUE), fb9, ac_w)
        cur += 1

        for idx, pl in enumerate(plans):
            ws.row_dimensions[cur].height = 40
            dos_raw = pl.get("dosParticipation") or ""
            _w(ws, cur, 1, idx + 1,                          _fill(C_GRAY_N), fb9, ac)
            _w(ws, cur, 2, pl.get("title", ""),               None, f9, al_w)
            _w(ws, cur, 3, pl.get("objectives", ""),          None, f9, al_w)
            _w(ws, cur, 4, pl.get("location", ""),            None, f9, al_w)
            _w(ws, cur, 5, pl.get("plannedDates", ""),        None, f9, ac_w)
            _w(ws, cur, 6, DOS_DISPLAY.get(dos_raw, dos_raw), None, f9, ac)
            _w(ws, cur, 7, pl.get("observations", ""),        None, f9, al_w)
            cur += 1

        ws.row_dimensions[cur].height = 8
        cur += 1

    if not has_c:
        ws.row_dimensions[cur].height = 20
        _mw(ws, cur, 1, 7, "(Aucune activité planifiée renseignée)",
            None, _font(italic=True, size=9), _align("center", "center"))
        cur += 1

    # ══════════════════════════════════════════════
    # SECTION D
    # ══════════════════════════════════════════════
    ws.row_dimensions[cur].height = 21
    _mw(ws, cur, 1, 7,
        "  D.  POINTS DE VIGILANCE  /  RISQUES  /  BESOINS DE DÉCISION",
        _fill(C_DARK_RED), fw10, _align("left", "center"))
    cur += 1

    d_hdrs = [
        "N°", "Catégorie", "Description du point",
        "Niveau de criticité", "Action attendue / Décision souhaitée", "", "",
    ]

    has_d = False
    for code in ENTITY_ORDER:
        risks = [
            r for r in all_data[code]["risks"]
            if r.get("description", "").strip() or r.get("category", "").strip()
        ]
        if not risks:
            continue
        has_d = True
        name = entity_names.get(code, code)

        ws.row_dimensions[cur].height = 16
        _mw(ws, cur, 1, 7, name, _fill(C_NOTE_D), fb9, ac)
        cur += 1

        ws.row_dimensions[cur].height = 26
        for i, txt in enumerate(d_hdrs[:5], start=1):
            _w(ws, cur, i, txt, _fill(C_PALE_BLUE), fb9, ac_w)
        cur += 1

        for idx, rk in enumerate(risks):
            ws.row_dimensions[cur].height = 40
            crit_raw = rk.get("criticality") or ""
            _w(ws, cur, 1, idx + 1,                                _fill(C_GRAY_N), fb9, ac)
            _w(ws, cur, 2, rk.get("category", ""),                 None, f9, al_w)
            _w(ws, cur, 3, rk.get("description", ""),              None, f9, al_w)
            _w(ws, cur, 4, CRITICALITY_DISPLAY.get(crit_raw, crit_raw), None, f9, ac)
            _w(ws, cur, 5, rk.get("expectedAction", ""),           None, f9, al_w)
            cur += 1

        ws.row_dimensions[cur].height = 8
        cur += 1

    if not has_d:
        ws.row_dimensions[cur].height = 20
        _mw(ws, cur, 1, 7, "(Aucun point de vigilance renseigné)",
            None, _font(italic=True, size=9), _align("center", "center"))
        cur += 1

    _apply_comp_borders(ws)


# ─── Utilitaires nom de fichier ────────────────────────────────────────────────

def make_output_filename(start: str, end: str) -> str:
    d0 = datetime.strptime(start, "%Y-%m-%d")
    d1 = datetime.strptime(end,   "%Y-%m-%d")
    m  = FRENCH_MONTHS[d1.month]
    return f"Bulletin_Hebdo_LHSPLA_Semaine_du_{d0.day}_au_{d1.day}_{m}_{d1.year}.xlsx"


def make_bulletin_tab_name(start: str, end: str) -> str:
    d0 = datetime.strptime(start, "%Y-%m-%d")
    d1 = datetime.strptime(end,   "%Y-%m-%d")
    m3 = FRENCH_MONTHS[d1.month][:3]
    return f"BulletinDu_{d0.day}-{d1.day}{m3}{d1.year}"


# ─── Point d'entrée ────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Génère le Bulletin Hebdomadaire LHSPLA-TA")
    ap.add_argument("--semaine",    help="Date de début de semaine YYYY-MM-DD (défaut : dernière)")
    ap.add_argument("--output-dir", default=SCRIPT_DIR, help="Répertoire de sortie")
    args = ap.parse_args()

    print("[1/6] Connexion a la base de donnees...")
    conn = db_connect()

    print("[2/6] Chargement de la semaine...")
    week       = load_week(conn, args.semaine)
    week_id    = week["id"]
    week_ref   = week["weekReference"]
    week_start = week["weekStart"]
    week_end   = week["weekEnd"]
    print(f"      -> {week_ref}  ({week_start} -> {week_end})")

    print("[3/6] Chargement des donnees entites...")
    all_data = load_entity_data(conn, week_id)
    conn.close()

    out_name = make_output_filename(week_start, week_end)
    out_path = os.path.join(args.output_dir, out_name)

    print(f"[4/6] Copie du fichier de reference -> {out_name}")
    shutil.copy(REFERENCE_FILE, out_path)

    print("[5/6] Chargement du classeur...")
    wb = load_workbook(out_path)

    # Extraction des noms d'entités depuis les onglets de référence (avant modification)
    entity_names: dict[str, str] = {}
    for code in ENTITY_ORDER:
        if code in wb.sheetnames:
            raw = wb[code].cell(row=2, column=1).value or code
            # Retirer le préfixe "Entité : " si présent
            name = str(raw).removeprefix("Entite : ").removeprefix("Entité : ").strip()
            entity_names[code] = name
        else:
            entity_names[code] = code
    print(f"      Entites: {', '.join(entity_names.keys())}")

    # ── Lecture des largeurs de colonnes de la référence avant reconstruction ─
    from openpyxl.utils import get_column_letter
    ref_col_widths: dict[str, dict] = {}
    for code in ENTITY_ORDER:
        if code in wb.sheetnames:
            ws_ref = wb[code]
            ref_col_widths[code] = {
                get_column_letter(c): ws_ref.column_dimensions[get_column_letter(c)].width
                for c in range(1, 9)
            }

    # ── Reconstruction des onglets entités ────────────────────────────────────
    print("[6/6] Reconstruction des onglets...")
    for code in ENTITY_ORDER:
        full = f"Entité : {entity_names[code]}"
        ws = _reset_sheet(wb, code)
        build_entity_sheet(ws, code, full, all_data[code], week_ref,
                           week_start=week_start, week_end=week_end,
                           col_widths=ref_col_widths.get(code))
        print(f"      OK {code}")

    # ── Reconstruction COMPILATION ────────────────────────────────────────────
    ws_comp = _reset_sheet(wb, "COMPILATION")
    build_compilation_sheet(ws_comp, week_ref, entity_names)
    print("      OK COMPILATION")

    # ── Reconstruction BulletinDu ─────────────────────────────────────────────
    bul_name = make_bulletin_tab_name(week_start, week_end)
    existing = next((n for n in wb.sheetnames if n.startswith("BulletinDu")), None)
    if existing:
        pos = wb.sheetnames.index(existing)
        del wb[existing]
        ws_bul = wb.create_sheet(bul_name, pos)
    else:
        pos = wb.sheetnames.index("COMPILATION") + 1 if "COMPILATION" in wb.sheetnames else len(wb.sheetnames)
        ws_bul = wb.create_sheet(bul_name, pos)

    build_bulletin_sheet(ws_bul, week_ref, entity_names, all_data)
    print(f"      OK {bul_name}")

    # ── Sauvegarde ───────────────────────────────────────────────────────────
    wb.save(out_path)
    print(f"\nFichier genere : {out_path}")


if __name__ == "__main__":
    main()
